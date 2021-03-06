# -*- coding: utf-8 -*-

#! python3

import openpyxl, sys

from openpyxl.styles import Font
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell

try:
	from openpyxl.cell import column_index_from_string,get_column_letter
except ImportError:
	from openpyxl.utils import column_index_from_string,get_column_letter

import logging
logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")
logging.disable(logging.CRITICAL)

wb = openpyxl.Workbook() # create the workbook
sheet = wb.active # switch to the active sheet, there should only be one
sheet.title = "Multi Table" # rename the sheet

# parse the command line

multi_number_cmdline = sys.argv[1] # should be the first argument in the command line i.e. the number, if you use 0 you get the program name

# multi_number_cmdline = "3" # set for purposes of testing, disable at end of development

logging.debug('The command line value entered for the max multi-table value is:  %s' % (multi_number_cmdline))

# create the font style for the headers

boldHeaderFont1 = Font(bold=True) # font object

# create the frozen header row

for x in range(2,int(multi_number_cmdline) + 2):  # since you started the loop at 2, then you need to shift the ending value by 2
	column_letter = get_column_letter(x)
	# the row stays the same i.e. 1
	logging.debug('Frozen row header - The current header column letter is:  %s' % (column_letter))
	sheet[column_letter + '1'] = x - 1 # set that cell to the current x value in the loop
	logging.debug('Frozen row header - The current header column letter and row number changed is:  %s' % (column_letter + '1'))
	# alter the font to bold
	sheet[column_letter + '1'].font = boldHeaderFont1
	logging.debug('The font for %s has been altered.' % (column_letter + '1'))

# create the frozen header column

for x in range(2,int(multi_number_cmdline) + 2):
	sheet["A" + str(x)] = x - 1
	logging.debug('Frozen column header - The current header column letter and row number changed is:  %s' % ("A" + str(x)))
	# alter the font to bold
	sheet["A" + str(x)].font = boldHeaderFont1
	logging.debug('The font for %s has been altered.' % ("A" + str(x)))

# freeze the panes

sheet.freeze_panes = 'B2' # row 1 and columns A
logging.debug('Panes frozen creating header row and header column.')

# begin filling in the multiplication values

for colValue in range(2,int(multi_number_cmdline) + 2):
	column_letter = get_column_letter(colValue)
	logging.debug('Frozen row header - The current header column letter is:  %s' % (column_letter))
	logging.debug('Frozen row header - The current header column label is:  %i' % (colValue - 1))

	# now to fill each value row by row within the column
	for rowValue in range(2,int(multi_number_cmdline) + 2):
		# logging.debug('The current column value is:  %i' % (colValue - 1))
		# logging.debug('The current row value is:  %i' % (rowValue - 1))
		sheet[column_letter + str(rowValue)] = (colValue - 1) * (rowValue - 1)
		logging.debug('The multiplication value set is:  %i' % ((colValue - 1) * (rowValue - 1)))

# save the final sheet

wb.save('multiplicationTable.xlsx')
logging.debug('Spreadsheet file saved.')
